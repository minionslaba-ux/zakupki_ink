"""Каталог: загрузка и индексация всех справочников калькулятора.

:class:`Catalog` держит в памяти сырьё, полупродукты, продукты, рецептуры,
курсы и потери, а также индексы для быстрого поиска по имени/артикулу.

Источники данных:

* :meth:`Catalog.from_json` — встроенный снимок (``data/nev_data.json``),
  извлечённый из ``Калькулятор_NEV_v9 (1).xlsx``. Не требует зависимостей;
* :meth:`Catalog.from_xlsx` — чтение «живой» книги Excel напрямую
  (требуется пакет ``openpyxl``).
"""

from __future__ import annotations

import json
from dataclasses import dataclass, field
from pathlib import Path

from .models import (
    LossRecord,
    Product,
    RawMaterial,
    Rate,
    RecipeLine,
    SemiProduct,
)

# Путь к встроенному снимку данных.
DEFAULT_DATA_PATH = Path(__file__).with_name("data") / "nev_data.json"

# Имена листов в исходной книге Excel.
SHEET_RAW = "Сырьё"
SHEET_SEMI = "Полупродукты"
SHEET_PRODUCTS = "Продукты"
SHEET_RECIPES = "Рецептуры"
SHEET_RATES = "Курсы ЦБ"
SHEET_LOSSES = "Потери"


def _norm(name: str | None) -> str:
    """Нормализовать имя для сопоставления (трим + схлопывание пробелов).

    В исходных данных встречаются двойные пробелы и хвостовые пробелы,
    поэтому при поиске компонент рецептур имена нормализуются.
    """
    if name is None:
        return ""
    return " ".join(str(name).split())


@dataclass
class Catalog:
    """Контейнер со всеми справочниками и индексами.

    Обычно создаётся через :meth:`from_json` или :meth:`from_xlsx`.
    """

    raw_materials: list[RawMaterial] = field(default_factory=list)
    semi_products: list[SemiProduct] = field(default_factory=list)
    products: list[Product] = field(default_factory=list)
    recipes: list[RecipeLine] = field(default_factory=list)
    rates: list[Rate] = field(default_factory=list)
    losses: list[LossRecord] = field(default_factory=list)
    vat_rate: float = 0.22

    # --- индексы (заполняются в reindex) -------------------------------
    _raw_by_name: dict[str, RawMaterial] = field(default_factory=dict, repr=False)
    _semi_by_name: dict[str, SemiProduct] = field(default_factory=dict, repr=False)
    _product_by_name: dict[str, Product] = field(default_factory=dict, repr=False)
    _by_article: dict[str, str] = field(default_factory=dict, repr=False)
    _recipe_by_key: dict[str, RecipeLine] = field(default_factory=dict, repr=False)
    _recipe_by_owner: dict[str, list[RecipeLine]] = field(
        default_factory=dict, repr=False
    )

    # ------------------------------------------------------------------
    # Построение индексов
    # ------------------------------------------------------------------
    def reindex(self) -> "Catalog":
        """Перестроить все индексы. Вызывать после изменения справочников."""
        self._raw_by_name = {_norm(m.name): m for m in self.raw_materials}
        self._semi_by_name = {_norm(s.name): s for s in self.semi_products}
        self._product_by_name = {_norm(p.name): p for p in self.products}

        self._by_article = {}
        for p in self.products:
            if p.article:
                self._by_article[str(p.article)] = p.name
        for s in self.semi_products:
            if s.article:
                self._by_article.setdefault(str(s.article), s.name)

        self._recipe_by_key = {}
        self._recipe_by_owner = {}
        for line in self.recipes:
            self._recipe_by_key[line.key] = line
            self._recipe_by_owner.setdefault(_norm(line.owner), []).append(line)
        for lines in self._recipe_by_owner.values():
            lines.sort(key=lambda ln: ln.index)
        return self

    # ------------------------------------------------------------------
    # Поиск
    # ------------------------------------------------------------------
    def raw(self, name: str) -> RawMaterial | None:
        """Найти сырьё по имени (с нормализацией пробелов)."""
        return self._raw_by_name.get(_norm(name))

    def semi(self, name: str) -> SemiProduct | None:
        """Найти полупродукт по имени."""
        return self._semi_by_name.get(_norm(name))

    def product(self, name: str) -> Product | None:
        """Найти готовый продукт по имени."""
        return self._product_by_name.get(_norm(name))

    def is_semi(self, name: str) -> bool:
        """Является ли компонента полупродуктом (а не сырьём)."""
        return _norm(name) in self._semi_by_name

    def recipe_lines(self, owner: str) -> list[RecipeLine]:
        """Строки рецептуры продукта/полупродукта, отсортированные по номеру."""
        return self._recipe_by_owner.get(_norm(owner), [])

    def name_by_article(self, article: str) -> str | None:
        """Имя продукта/полупродукта по артикулу (поиск как в ячейке G8)."""
        return self._by_article.get(str(article))

    # ------------------------------------------------------------------
    # Загрузка из встроенного JSON-снимка
    # ------------------------------------------------------------------
    @classmethod
    def from_json(cls, path: str | Path | None = None) -> "Catalog":
        """Загрузить каталог из JSON-снимка.

        :param path: путь к JSON; по умолчанию — встроенный
            ``data/nev_data.json``.
        """
        path = Path(path) if path else DEFAULT_DATA_PATH
        data = json.loads(Path(path).read_text(encoding="utf-8"))

        cat = cls(vat_rate=float(data.get("meta", {}).get("vat_rate", 0.22)))

        cat.raw_materials = [
            RawMaterial(
                name=r["name"],
                price_eur=_safe_float(r.get("price_eur")),
                currency=r.get("currency") or "EUR",
                price_ccy=r.get("price_ccy"),
                date_serial=r.get("date"),
                unit=r.get("unit"),
                supplier=r.get("supplier"),
                article=r.get("article"),
            )
            for r in data.get("raw_materials", [])
        ]
        cat.semi_products = [
            SemiProduct(
                name=s["name"],
                price_eur=_safe_float(s.get("price_eur_stored")),
                n_comp=s.get("n_comp"),
                article=s.get("article"),
            )
            for s in data.get("semi_products", [])
        ]
        cat.products = [
            Product(
                name=p["name"],
                client=p.get("client"),
                type=p.get("type"),
                article=p.get("article"),
            )
            for p in data.get("products", [])
        ]
        cat.recipes = [
            RecipeLine(
                owner=r["product"],
                index=int(r["idx"]),
                component=r["component"],
                pct_base=_safe_float(r.get("pct_base")),
                pct_real=r.get("pct_real"),
                article=r.get("article"),
            )
            for r in data.get("recipes", [])
            if r.get("component")
        ]
        cat.rates = [
            Rate(
                date_serial=_safe_float(r["date_serial"]),
                eur_rub=_safe_float(r.get("eur_rub")),
                usd_rub=_safe_float(r.get("usd_rub")),
                cny_rub=_safe_float(r.get("cny_rub")),
                eur_usd=_safe_float(r.get("eur_usd")),
                eur_cny=_safe_float(r.get("eur_cny")),
                source=r.get("source"),
            )
            for r in data.get("rates", [])
            if r.get("date_serial") is not None
        ]
        cat.losses = [
            LossRecord(
                product=l["product"],
                component=l["component"],
                index=l.get("idx"),
                use_per_kg=l.get("use_per_kg"),
                sum_comp=l.get("sum_comp"),
                loss_kg=l.get("loss_kg"),
                batches=l.get("batches"),
                article=l.get("article"),
                key=l.get("key"),
            )
            for l in data.get("losses", [])
        ]
        return cat.reindex()

    # ------------------------------------------------------------------
    # Загрузка из «живой» книги Excel
    # ------------------------------------------------------------------
    @classmethod
    def from_xlsx(cls, path: str | Path) -> "Catalog":
        """Загрузить каталог напрямую из файла ``Калькулятор_NEV*.xlsx``.

        Требуется ``openpyxl`` (см. ``requirements.txt``). Читаются
        вычисленные значения ячеек (``data_only=True``), поэтому файл
        должен быть хотя бы раз пересчитан и сохранён Excel/Google Sheets.
        """
        try:
            from openpyxl import load_workbook
        except ModuleNotFoundError as exc:  # pragma: no cover
            raise RuntimeError(
                "Для чтения .xlsx требуется openpyxl: pip install openpyxl"
            ) from exc

        wb = load_workbook(path, data_only=True, read_only=True)

        def cell(ws, col: int, row: int):
            return ws.cell(row=row, column=col).value

        cat = cls()

        # --- Сырьё: C=name D=unit E=price€ F=date G=supplier H=art J=cur K=ccy
        ws = wb[SHEET_RAW]
        for row in range(6, ws.max_row + 1):
            name = cell(ws, 3, row)
            if not name:
                continue
            cat.raw_materials.append(
                RawMaterial(
                    name=str(name),
                    unit=cell(ws, 4, row),
                    price_eur=float(cell(ws, 5, row) or 0.0),
                    date_serial=_serial(cell(ws, 6, row)),
                    supplier=cell(ws, 7, row),
                    article=cell(ws, 8, row),
                    currency=cell(ws, 10, row) or "EUR",
                    price_ccy=cell(ws, 11, row),
                )
            )

        # --- Полупродукты: C=name D=n E=price€ F=art
        ws = wb[SHEET_SEMI]
        for row in range(6, ws.max_row + 1):
            name = cell(ws, 3, row)
            if not name:
                continue
            cat.semi_products.append(
                SemiProduct(
                    name=str(name),
                    n_comp=_int(cell(ws, 4, row)),
                    price_eur=float(cell(ws, 5, row) or 0.0),
                    article=cell(ws, 6, row),
                )
            )

        # --- Продукты: C=client D=name E=type F=art
        ws = wb[SHEET_PRODUCTS]
        for row in range(6, ws.max_row + 1):
            name = cell(ws, 4, row)
            if not name:
                continue
            cat.products.append(
                Product(
                    name=str(name),
                    client=cell(ws, 3, row),
                    type=cell(ws, 5, row),
                    article=cell(ws, 6, row),
                )
            )

        # --- Рецептуры: C=client D=product E=idx F=comp G=%base H=%real L=art
        ws = wb[SHEET_RECIPES]
        for row in range(6, ws.max_row + 1):
            owner = cell(ws, 4, row)
            comp = cell(ws, 6, row)
            idx = cell(ws, 5, row)
            if not owner or not comp or idx is None:
                continue
            cat.recipes.append(
                RecipeLine(
                    owner=str(owner),
                    index=int(idx),
                    component=str(comp),
                    pct_base=float(cell(ws, 7, row) or 0.0),
                    pct_real=cell(ws, 8, row),
                    article=cell(ws, 12, row),
                )
            )

        # --- Курсы ЦБ: B=date C=eurrub D=usdrub E=cnyrub F=eurusd G=eurcny H=src
        ws = wb[SHEET_RATES]
        for row in range(7, ws.max_row + 1):
            d = cell(ws, 2, row)
            if d is None:
                continue
            cat.rates.append(
                Rate(
                    date_serial=_serial(d),
                    eur_rub=float(cell(ws, 3, row) or 0.0),
                    usd_rub=float(cell(ws, 4, row) or 0.0),
                    cny_rub=float(cell(ws, 5, row) or 0.0),
                    eur_usd=float(cell(ws, 6, row) or 0.0),
                    eur_cny=float(cell(ws, 7, row) or 0.0),
                    source=cell(ws, 8, row),
                )
            )

        # --- Потери: B=art C=product D=idx E=comp F=use G=sum H=loss I=batch J=key
        if SHEET_LOSSES in wb.sheetnames:
            ws = wb[SHEET_LOSSES]
            for row in range(6, ws.max_row + 1):
                key = cell(ws, 10, row)
                if not key:
                    continue
                cat.losses.append(
                    LossRecord(
                        article=cell(ws, 2, row),
                        product=cell(ws, 3, row),
                        index=_int(cell(ws, 4, row)),
                        component=cell(ws, 5, row),
                        use_per_kg=cell(ws, 6, row),
                        sum_comp=cell(ws, 7, row),
                        loss_kg=cell(ws, 8, row),
                        batches=_int(cell(ws, 9, row)),
                        key=key,
                    )
                )

        wb.close()
        return cat.reindex()


def _serial(value) -> float | None:
    """Привести значение даты к серийному номеру Excel."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    # openpyxl возвращает datetime — переведём в серийный номер.
    from datetime import date, datetime

    from .models import date_to_excel_serial

    if isinstance(value, datetime):
        value = value.date()
    if isinstance(value, date):
        return float(date_to_excel_serial(value))
    return None


def _safe_float(value, default: float = 0.0) -> float:
    """Привести значение к float, вернуть ``default`` при ошибке/None."""
    if value is None or value == "":
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _int(value) -> int | None:
    """Безопасно привести к int (значения часто хранятся как float)."""
    if value is None:
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None
